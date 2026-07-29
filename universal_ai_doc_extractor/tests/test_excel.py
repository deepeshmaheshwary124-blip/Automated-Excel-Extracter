"""Tests for Excel service."""

import os
import tempfile
from pathlib import Path

import pytest
import openpyxl

from services.excel_service import ExcelService


class TestExcelService:
    def setup_method(self):
        self.service = ExcelService()
        self.tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self.test_path = self.tmp.name
        self.tmp.close()

    def teardown_method(self):
        Path(self.test_path).unlink(missing_ok=True)

    def test_create_workbook(self):
        self.service.create_workbook(self.test_path, ["Date", "Description", "Amount"])
        wb = openpyxl.load_workbook(self.test_path)
        assert "Data" in wb.sheetnames

        ws = wb["Data"]
        assert ws.cell(1, 1).value == "Date"
        assert ws.cell(1, 2).value == "Description"
        assert ws.cell(1, 3).value == "Amount"
        wb.close()

    def test_append_data(self):
        self.service.create_workbook(self.test_path, ["Date", "Description", "Amount"])
        self.service.append_data(self.test_path, "Data", {
            "Date": "2024-01-15",
            "Description": "Test entry",
            "Amount": 100.50,
        })

        wb = openpyxl.load_workbook(self.test_path)
        ws = wb["Data"]
        assert ws.cell(2, 1).value == "2024-01-15"
        assert ws.cell(2, 2).value == "Test entry"
        assert ws.cell(2, 3).value == 100.50
        wb.close()

    def test_validate_workbook(self):
        self.service.create_workbook(self.test_path)
        result = self.service.validate_workbook(self.test_path)
        assert result["valid"] is True
        assert "Data" in result["info"]["sheets"]

    def test_get_sheet_names(self):
        self.service.create_workbook(self.test_path)
        names = self.service.get_sheet_names(self.test_path)
        assert "Data" in names

    def test_get_next_available_row(self):
        self.service.create_workbook(self.test_path, ["A", "B"])
        next_row = self.service.get_next_available_row(self.test_path, "Data")
        assert next_row == 2

    def test_detect_best_sheet(self):
        self.service.create_workbook(self.test_path)
        sheet = self.service.detect_best_sheet(self.test_path)
        assert sheet == "Data"

    def test_backup_workbook(self):
        self.service.create_workbook(self.test_path)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            backup_path = f.name

        try:
            self.service.backup_workbook(self.test_path, backup_path)
            assert os.path.exists(backup_path)
        finally:
            Path(backup_path).unlink(missing_ok=True)
