"""Excel workbook handling service with style preservation."""

import logging
import shutil
import tempfile
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.enums import DocumentType


logger = logging.getLogger(__name__)


class ExcelService:
    def open_workbook(self, file_path: str, data_only: bool = False) -> openpyxl.Workbook:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {file_path}")
        return openpyxl.load_workbook(str(path), data_only=data_only)

    def save_workbook(self, workbook: openpyxl.Workbook, file_path: str) -> None:
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(path))
        logger.info("Workbook saved: %s", file_path)

    def get_sheet_names(self, file_path: str) -> list[str]:
        wb = self.open_workbook(file_path)
        names = wb.sheetnames
        wb.close()
        return names

    def get_next_available_row(self, file_path: str, sheet_name: str) -> int:
        wb = self.open_workbook(file_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        next_row = (ws.max_row or 0) + 1
        wb.close()
        return next_row

    def find_sheet_by_name(self, file_path: str, name: str) -> Optional[str]:
        names = self.get_sheet_names(file_path)
        for n in names:
            if name.lower() in n.lower():
                return n
        return None

    def detect_best_sheet(self, file_path: str, document_type: str = "") -> str:
        names = self.get_sheet_names(file_path)
        if not names:
            raise ValueError("Workbook has no sheets")

        if document_type:
            keywords = {
                "invoice": ["invoice", "inv"],
                "receipt": ["receipt", "rcpt"],
                "purchase_order": ["po", "purchase"],
                "bank_statement": ["statement", "bank"],
                "bill": ["bill", "ap", "payable"],
            }
            doc_lower = document_type.lower()
            for doc_type, kw_list in keywords.items():
                if doc_lower in doc_type or doc_type in doc_lower:
                    for name in names:
                        if any(kw in name.lower() for kw in kw_list):
                            return name

        if "data" in names:
            return "data"
        if "Sheet1" in names:
            return "Sheet1"

        return names[0]

    def append_data(self, file_path: str, sheet_name: str,
                    data: dict[str, Any], headers: Optional[list[str]] = None) -> dict[str, Any]:
        wb = self.open_workbook(file_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        next_row = (ws.max_row or 0) + 1

        if next_row == 1 and headers:
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            next_row = 2

        for col, (key, value) in enumerate(data.items(), 1):
            cell = ws.cell(row=next_row, column=col, value=value)
            if next_row > 1:
                above_cell = ws.cell(row=next_row - 1, column=col)
                if above_cell.font:
                    try:
                        from copy import copy
                        cell.font = copy(above_cell.font)
                        cell.alignment = copy(above_cell.alignment)
                        cell.border = copy(above_cell.border)
                        cell.fill = copy(above_cell.fill)
                        cell.number_format = above_cell.number_format
                    except Exception:
                        pass

        tmp_path = str(Path(file_path).with_suffix(".tmp.xlsx"))
        try:
            wb.save(tmp_path)
            shutil.move(tmp_path, file_path)
        except Exception as e:
            Path(tmp_path).unlink(missing_ok=True)
            raise

        wb.close()

        return {"row": next_row, "columns": list(data.keys())}

    def create_workbook(self, file_path: str, headers: Optional[list[str]] = None) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)

        wb.save(str(file_path))
        wb.close()
        logger.info("Workbook created: %s", file_path)
        return file_path

    def validate_workbook(self, file_path: str) -> dict[str, Any]:
        result = {"valid": True, "errors": [], "warnings": [], "info": {}}
        try:
            wb = self.open_workbook(file_path)
            result["info"]["sheets"] = wb.sheetnames
            result["info"]["sheet_count"] = len(wb.sheetnames)

            for name in wb.sheetnames:
                ws = wb[name]
                result["info"][f"sheet_{name}_rows"] = ws.max_row
                result["info"][f"sheet_{name}_cols"] = ws.max_column

            if wb.security and wb.security.lockStructure:
                result["warnings"].append("Workbook structure is locked")

            wb.close()
        except Exception as e:
            result["valid"] = False
            result["errors"].append(str(e))

        return result

    def backup_workbook(self, file_path: str, backup_path: str) -> str:
        shutil.copy2(file_path, backup_path)
        logger.info("Workbook backed up: %s -> %s", file_path, backup_path)
        return backup_path
